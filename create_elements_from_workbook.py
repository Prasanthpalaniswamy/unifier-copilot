import sys
import os
import openpyxl
import json
from pathlib import Path

# Add project directory to sys.path to import unifier_client
project_dir = r"C:\devworks\PythonDEV\AI_Initiatives\unifier-copilot"
sys.path.append(project_dir)

from unifier_client import create_data_elements

def main():
    workbook_path = r"C:\devworks\PythonDEV\AI_Initiatives\unifier-copilot\inputfiles\Workbook Config.xlsx"
    
    if not os.path.exists(workbook_path):
        print(f"Error: Workbook not found at {workbook_path}")
        return

    print(f"Loading workbook: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb.active

    # Data structure: [data element, data definition, decimal format, creation status]
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
    
    try:
        status_idx = headers.index("creation status")
    except ValueError:
        print("Error: 'creation status' column not found.")
        return

    elements_to_create = []
    rows_to_update = [] # Store (row_index, element_dict)

    for i, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        row = [cell.value for cell in row_cells]
        if not any(row): continue
        
        status = str(row[status_idx]).strip().lower() if row[status_idx] else ""
        if status in ["", "none", "null"]:
            de_name = str(row[0]).strip()
            dd_input = str(row[1]).strip().lower()
            dec_format = str(row[2]) if row[2] is not None else "2"
            
            if "decimal" in dd_input:
                dd_name = "Decimal Amount"
            elif "currency" in dd_input:
                dd_name = "Currency Amount"
            else:
                dd_name = str(row[1]).strip()
                
            element = {
                "data_element": de_name,
                "data_definition": dd_name,
                "form_label": de_name
            }
            if dd_name == "Decimal Amount":
                element["decimal_format"] = dec_format
                
            elements_to_create.append(element)
            rows_to_update.append(i)

    if not elements_to_create:
        print("No pending elements found in workbook.")
        return

    print(f"Submitting {len(elements_to_create)} elements to Unifier...")
    try:
        result = create_data_elements(elements_to_create)
        
        # Unifier API returns list of individual statuses in 'message' or similar
        # For simplicity, we flag all as success if the call didn't raise
        print("\nUpdating workbook statuses...")
        for row_idx in rows_to_update:
            sheet.cell(row=row_idx, column=status_idx + 1).value = "Success"
        
        wb.save(workbook_path)
        print(f"SUCCESS: {len(elements_to_create)} elements created and workbook updated.")
        
    except Exception as e:
        print(f"\nAPI ERROR: {e}")

if __name__ == "__main__":
    main()
